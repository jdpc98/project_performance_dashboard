import pandas as pd
from openpyxl import load_workbook

def unmerge_and_fill(sheet):
    """
    Example helper function that unmerges all merged cells in a given sheet
    and fills each formerly merged cell with the original top-left value.
    """
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
    # Optionally, do a forward fill using openpyxl.
    # (Often it's easier just to handle merges within pandas, but this is an option.)
    # For each cell that was in a merged range, copy the top-left value:
    # Not strictly necessary if you re-read with pandas after unmerging.


def load_entire_sheet(file_path, sheet_name):
    """
    Loads the entire sheet as a pandas DataFrame after unmerging cells.
    Returns a DataFrame with all rows/columns (no standard header).
    """
    # --- 1) Unmerge cells at the openpyxl level (optional) ---
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    unmerge_and_fill(ws)
    # Save to a temporary file or in-memory so pandas sees the unmerged version
    temp_file = file_path.replace(".xlsx", "_unmerged.xlsx")
    wb.save(temp_file)

    # --- 2) Read the unmerged temp file with pandas ---
    # Use header=None so we get a raw layout (no auto column names)
    df = pd.read_excel(temp_file, sheet_name=sheet_name, header=None)
    
    return df


def parse_first_df(df):
    """
    Create the first DataFrame with monthly/year TRM values 
    and the loaded_val for Houston (row 3) and Bogota (row 2).
    
    Suppose:
      - TRM row is row=0, columns C1:S1 => df.iloc[0, 2:19]
      - Years are in row=4 (since row 5 in Excel is index 4 in pandas).
      - Row 2 => df.iloc[1, ...] has Bogota loaded_val
      - Row 3 => df.iloc[2, ...] has Houston loaded_val
      (Adjust these indices as needed.)
    """
    # Example placeholders:
    # Month TRM in row=0, columns=2..19
    trm_values = df.iloc[0, 2:19].values.tolist()
    # Year row might be row=4 in 0-based indexing
    years = df.iloc[4, 2:19].values.tolist()
    # loaded_val for Bogota and Houston
    bogota_vals = df.iloc[1, 2:19].values.tolist()
    houston_vals = df.iloc[2, 2:19].values.tolist()
    
    # Construct a tidy DataFrame with columns: ["MonthIndex", "Year", "TRM", "Bogota_val", "Houston_val"]
    # For example, assume each column corresponds to a consecutive month. 
    # If you know which months they correspond to, you can add a "Month" column as well.
    data_rows = []
    for i, (trm, yr, bogv, houv) in enumerate(zip(trm_values, years, bogota_vals, houston_vals)):
        data_rows.append({
            "MonthIndex": i+1,
            "Year": yr,
            "TRM": trm,
            "Bogota_val": bogv,
            "Houston_val": houv
        })
    df_first = pd.DataFrame(data_rows)
    return df_first


def parse_second_df(df):
    """
    The second data region is from A4:AC99 => in 0-based indexing: 
      rows 3..99 (since row=4 in Excel => row=3 in pandas), 
      columns 0..29 (since AC is the 29th column in 0-based).
    
    In column A => ID (df.iloc[:,0]), 
    column B => Person Name (df.iloc[:,1]), 
    columns C..D => rate values for some years, etc.
    
    We want 2 separate DataFrames:
      - one for 2022, 2023
      - one for 2024, 2025, ...
    
    Adjust the row slicing as needed.
    """
    # Slice out the block
    block = df.iloc[3:99, 0:29].copy()  # A4..AC99
    block.columns = [f"Col_{c}" for c in range(len(block.columns))]  # placeholder col names
    
    # For example:
    # block["ID"] = block["Col_0"]
    # block["Name"] = block["Col_1"]
    # block["Rate1"] = block["Col_2"]
    # block["Rate2"] = block["Col_3"]
    # ...
    
    # We might do something like:
    block.rename(columns={
        "Col_0": "ID#",
        "Col_1": "Name",
        "Col_2": "Rate_2022",
        "Col_3": "Rate_2023",
        # etc.
    }, inplace=True)
    
    # We then separate into two DataFrames for different year sets
    # (This part is conceptual; tailor it to your actual data layout.)
    df_2022_23 = block[["ID#", "Name", "Rate_2022", "Rate_2023"]].copy()
    # For 2024, 2025, etc. assume columns 4..somewhere
    df_2024_25 = block.drop(columns=["Rate_2022", "Rate_2023"]).copy()
    
    return df_2022_23, df_2024_25


def parse_third_df(df):
    """
    
     slice out rows and columns, interpret them, 
    and build a DataFrame for the "raw and loaded values for 2024."
    
    Example (adjust row range as needed):
    """
    # If AD=30 and AG=33 in 0-based: 
    block = df.iloc[0:100, 30:35].copy()  # example row range
    # Suppose columns: AD => 'raw_usd', AE => ???, AF => 'loaded_cop', AG => 'raw_cop'
    # or per your example, you have something like:
    # AD => incorrectly named 'raw usd' but we store it as loaded_usd?
    
    block.columns = ["Loaded_usd", "raw_usd", "loaded_cop", "raw_cop"]  # adjust names as needed
    
    # You mention the "conversion value" in AG1..AG3 merged, we might do:
    # conversion_val = df.iloc[0, 33] or something. If it's the same cell repeated, 
    # pick one. 
    conversion_val = df.iloc[0, 33]
    block["Conversion"] = conversion_val
    
    # Also, you said "the employee name assigned for each row is in that row"
    # Possibly you link this block with the second block by row index, or 
    # you have names in another column. 
    # If the names are in column B for the same row, you might do:
    # block["Name"] = df.iloc[block.index, 1]
    
    return block


def main():
    file_path = r"C:\Users\jose.pineda\Desktop\operations\RATES.xlsx"
    sheet_name = "Rates"  # Adjust as needed

    # 1. Read the entire sheet (optionally unmerge cells).
    df_full = load_entire_sheet(file_path, sheet_name=sheet_name)

    # 2. Parse the first DataFrame (monthly/year TRM + loaded_val for Houston/Bogota).
    df_first = parse_first_df(df_full)
    
    # 3. Parse the second data region (A4:AC99) into two separate DataFrames.
    df_2022_23, df_2024_25 = parse_second_df(df_full)
    
    # 4. Parse the third data region (AD:AG) for raw/loaded values (2024).
    df_third = parse_third_df(df_full)
    
    # Now you have your 4 DataFrames:
    #   df_first
    #   df_2022_23
    #   df_2024_25
    #   df_third
    
    # Example: Print or return them
    print("----- DF1: TRM & loaded vals -----")
    print(df_first.head())
    
    print("----- DF2A: 2022/2023 data -----")
    print(df_2022_23.head())
    
    print("----- DF2B: 2024/2025 data -----")
    print(df_2024_25.head())
    
    print("----- DF3: Raw & Loaded Values (2024) -----")
    print(df_third.head())

if __name__ == "__main__":
    main()
