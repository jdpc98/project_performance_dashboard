import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import warnings

# Dash and Plotly imports
import dash
from dash import dcc, html, dash_table
import plotly.express as px

# Optionally suppress UserWarnings (e.g., from openpyxl)
warnings.simplefilter("ignore", UserWarning)

def process_files():
    """
    Prompts the user to select File1 (destination) and File2 (source),
    backs up the "Staff" sheet from File1, processes File2 (filtering rows
    based on the jobcode_1 column and dropping that column), updates File1,
    calculates hours (local_end_time - local_start_time), extracts the project
    code (first 7 characters of jobcode_2), and returns the processed DataFrame.
    """
    # Hide the Tkinter root window
    root = tk.Tk()
    root.withdraw()

    # --- STEP 1: Select the destination file (File1) ---
    file1_path = filedialog.askopenfilename(
        title="Select the destination file (File1)",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xls")]
    )
    if not file1_path:
        print("No destination file selected. Exiting.")
        return None

    # --- STEP 2: Backup the current "Staff" sheet from File1 ---
    try:
        df_backup = pd.read_excel(file1_path, sheet_name="Staff")
    except Exception as e:
        print("Error reading 'Staff' sheet from destination file:", e)
        return None

    backup_filename = f"StaffBackup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    backup_path = os.path.join(os.path.dirname(file1_path), backup_filename)
    df_backup.to_excel(backup_path, index=False)
    print("Backup saved to:", backup_path)

    # --- STEP 3: Select the source file (File2) ---
     
    
    rates_file_path = filedialog.askopenfilename(
        title="Select the rates file (File2)",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xls")]
    )
    if not rates_file_path:
        print("No source file selected. Exiting.")
        return None

    # --- STEP 4: Read File2 into a DataFrame ---
    try:
        df_source = pd.read_excel(rates_file_path, sheet_name="Rates")
    except Exception as e:
        print("Error reading source file:", e)
        return None

    # Use forward fill to simulate unmerging of merged cells
    df_source.ffill(inplace=True)

    # --- STEP 5: Filter based on the "jobcode_1" column ---
    if 'ID#' not in df_source.columns:
        print("The source file does not contain a 'ID#' column.")
        return None

    #excluded = ["Lunch Break", "OVERHEAD", "Solerium", "Viu Group"]
    #df_filtered = df_source[~df_source['jobcode_1'].isin(excluded)]

    # Drop the 'jobcode_1' column so that the remaining columns match your destination
    df_final = df_filtered.drop(columns=['jobcode_1'])
    
    # --- NEW STEP: Calculate hours and extract project code ---
    # Convert the time columns to datetime objects.
    df_final['local_start_time'] = pd.to_datetime(df_final['local_start_time'], errors='coerce')
    df_final['local_end_time']   = pd.to_datetime(df_final['local_end_time'], errors='coerce')
    
    # Calculate the difference in hours
    df_final['hours'] = (df_final['local_end_time'] - df_final['local_start_time']).dt.total_seconds() / 3600.0

    # Extract project code: first 7 characters of jobcode_2
    df_final['project'] = df_final['jobcode_2'].astype(str).str[:7]
    # --- End NEW STEP ---

    # --- STEP 6: Update File1's "Staff" sheet using openpyxl ---
    wb = load_workbook(file1_path)
    if "Staff" not in wb.sheetnames:
        print("Destination file does not contain a sheet named 'Staff'.")
        return None
    ws = wb["Staff"]

    # Clear the destination range (columns D to AA, i.e., columns 4 to 27, starting at row 2)
    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, min_col=4, max_row=max_row, max_col=27):
        for cell in row:
            cell.value = None

    # Write new data from df_final into the "Staff" sheet starting at cell D2
    start_row = 2
    start_col = 4  # Column D
    for r_idx, row in enumerate(df_final.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file1_path)
    print("Data replaced successfully in the 'Staff' sheet.")

    # --- STEP 7: Return the processed DataFrame ---
    # We assume that the processed data now contains at least 'username', 'hours', and 'project'
    if 'username' not in df_final.columns or 'hours' not in df_final.columns:
        print("Data does not contain 'username' and/or 'hours' columns for further processing.")
        return None

    return df_final

def create_dash_app(df):
    """
    Creates and returns a Dash app that displays, for each project found in the
    processed DataFrame, a pie chart (grouped by username with total hours)
    and a table of details (username and hours), laid out side by side.
    """
    # Get the unique project codes
    projects = df['project'].unique()
    project_components = []

    # For each project, group data by username and sum hours
    for proj in projects:
        df_proj = df[df['project'] == proj]
        df_group = df_proj.groupby('username', as_index=False)['hours'].sum()
        
        # Create a pie chart for the project
        fig = px.pie(
            df_group,
            values='hours',
            names='username',
            title=f"Project: {proj}",
            hole=0.3
        )
        
        # Create a table (using dash_table) showing username and hours
        table = dash_table.DataTable(
            data=df_group.to_dict('records'),
            columns=[
                {'name': 'Username', 'id': 'username'},
                {'name': 'Hours Worked', 'id': 'hours'}
            ],
            style_cell={'textAlign': 'center', 'padding': '5px'},
            style_header={'fontWeight': 'bold'},
            page_size=10
        )
        
        # Create a Div with a two-column layout: left for the pie chart, right for the table
        proj_div = html.Div([
            html.Div(dcc.Graph(figure=fig), style={'flex': '1'}),
            html.Div(table, style={'flex': '1', 'padding': '20px'})
        ], style={
            'display': 'flex',
            'flexDirection': 'row',
            'alignItems': 'center',
            'marginBottom': '40px'
        })
        
        project_components.append(proj_div)
    
    # Define the overall layout
    app = dash.Dash(__name__)
    app.layout = html.Div([
        html.H1("Project Hours Analysis", style={'textAlign': 'center', 'marginBottom': '40px'}),
        *project_components  # Unpack the list of project components into the layout
    ], style={'width': '90%', 'margin': '0 auto'})
    
    return app

def main():
    # Process the files and get the processed DataFrame
    df_processed = process_files()
    if df_processed is None:
        print("Processing failed. Exiting.")
        return

    # Create the Dash app using the processed DataFrame
    app = create_dash_app(df_processed)

    # Run the Dash server on your specified host and port
    app.run_server(host='10.1.2.154', port=9050, debug=True)

if __name__ == '__main__':
    main()
